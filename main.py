from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.textinput import TextInput
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.spinner import Spinner
from kivy.uix.scrollview import ScrollView
from kivy.uix.popup import Popup
from kivy.uix.gridlayout import GridLayout
from kivy.core.window import Window
from kivy.utils import platform
from kivy.clock import Clock

from openpyxl import Workbook, load_workbook
import os
import json
import time
import urllib.parse
import webbrowser
from collections import Counter

IP server push ke liye network library
import threading
try:
import requests
except ImportError:
requests = None

class ScannerApp(App):

def build(self):
self.scan_data = [] # Memory mein live data rakhne ke liye list
self.duplicates = set() # Duplicate check karne ke liye set

# --- SETTINGS VARIABLES ---
self.selected_location = "Documents" # Default folder
self.selected_file_type = ".xlsx" # Default file type
self.selected_ip = "192.168.1.50" # Default Plant Central PC IP
self.selected_port = "5000" # Default Server Port
self.dispatch_location = "Rolling Mill-1" # Default Target Dispatch Location
self.enable_ip_push = True # Live IP server push hamesha default ON

# Live Key Tracking for Scan vs Manual detection
self.last_input_time = 0
self.is_manual_entry = False

# --- REPORT MEMORY STORAGE FOR EDITS ---
self.rep_cust_name = "Highway Indust."
self.rep_front_discard = "750"
self.rep_end_discard = "850"
self.rep_cooling_type = "Fast Fan Cooling"
self.rep_source_mode = "1. Use Current Live Scan"

# Popup open tracker flag taaki background focus block rahe
self.is_popup_open = False

self.update_storage_path()
self.load_dispatch_locations_from_storage()

# --- MAIN UI LAYOUT ---
main_layout = BoxLayout(orientation="vertical", padding=15, spacing=10)

# --- TOP ROW: HEAT INPUT & SETTING BUTTON ---
top_row = BoxLayout(orientation="horizontal", size_hint=(1, 0.1), spacing=10)

self.filename = TextInput(
hint_text="Heat File Name (Locked until Scan)",
multiline=False,
size_hint=(0.8, 1),
font_size='16sp',
readonly=True
 )
top_row.add_widget(self.filename)

self.settings_btn = Button(
text="⚙️",
size_hint=(0.2, 1),
background_color=(0.3, 0.3, 0.3, 1),
font_size='22sp'
)
self.settings_btn.bind(on_release=self.open_settings_popup)
top_row.add_widget(self.settings_btn)

main_layout.add_widget(top_row)

# --- ROW: HALF-WIDTH SPLIT BUTTONS ---
button_row = BoxLayout(orientation="horizontal", size_hint=(1, 0.08), spacing=10)

self.open_btn = Button(
text="📂 Open File",
size_hint=(0.5, 1),
background_color=(0.2, 0.6, 0.8, 1),
font_size='15sp'
)
self.open_btn.bind(on_press=self.open_existing_file)
button_row.add_widget(self.open_btn)

self.clear_btn = Button(
text="🧹 Clear Heat",
size_hint=(0.5, 1),
background_color=(0.7, 0.2, 0.2, 1),
font_size='15sp'
)
self.clear_btn.bind(on_press=self.clear_active_heat_data)
button_row.add_widget(self.clear_btn)

main_layout.add_widget(button_row)

# --- MAIN MASTER INPUT BOX ---
main_layout.add_widget(Label(text="[b]Bar Code Scan & Paste Center[/b]", markup=True, size_hint=(1, 0.04), font_size='14sp'))

self.barcode = TextInput(
hint_text="Just PASTE Data Here (No Enter Needed)...",
multiline=False,
size_hint=(1, 0.12),
font_size='16sp'
)
self.barcode.bind(text=self.on_text_change)
main_layout.add_widget(self.barcode)

# --- CENTER DASHBOARD BOX ---
center_box = BoxLayout(orientation="vertical", padding=10, spacing=5, size_hint=(1, 0.48))

self.count_label = Label(
text="No. of Count: 0", font_size='24sp', bold=True, color=(0, 1, 0, 1), size_hint=(1, 0.18)
)
center_box.add_widget(self.count_label)

# Last Scan and List Header
header_row = BoxLayout(orientation="horizontal", size_hint=(1, 0.1), spacing=10)
self.last_label = Label(
text="Last: None", font_size='12sp', halign='left', valign='middle', text_size=(Window.width * 0.5, None)
)
self.list_title = Label(
text="Last 10 Count", font_size='13sp', bold=True, halign='right', valign='middle', text_size=(Window.width * 0.4, None)
)
header_row.add_widget(self.last_label)
header_row.add_widget(self.list_title)
center_box.add_widget(header_row)

# ScrollView for History Records
scroll_view = ScrollView(size_hint=(1, 0.36), do_scroll_x=False, do_scroll_y=True)
self.last10_container = BoxLayout(orientation="vertical", size_hint_y=None, spacing=3)
self.last10_container.bind(minimum_height=self.last10_container.setter('height'))

self.history_labels = []
for _ in range(10):
lbl = Label(text="", font_size='13sp', size_hint_y=None, height=22, halign='center')
self.last10_container.add_widget(lbl)
self.history_labels.append(lbl)
scroll_view.add_widget(self.last10_container)
center_box.add_widget(scroll_view)

# EMERGENCY CONTROL ROW: Undo Last Scan
self.undo_btn = Button(
text="↩️ Undo Last Scan", size_hint=(1, 0.14), background_color=(0.9, 0.5, 0.1, 1), font_size='14sp', bold=True
)
self.undo_btn.bind(on_press=self.undo_last_scan_row)
center_box.add_widget(self.undo_btn)

# Generate Production Report
self.report_btn = Button(
text="📊 Generate Production Report",
size_hint=(1, 0.16),
background_color=(0.1, 0.5, 0.6, 1),
font_size='15sp',
bold=True
)
self.report_btn.bind(on_press=self.open_production_report_popup)
center_box.add_widget(self.report_btn)

# Save Button
self.save_btn = Button(
text="💾 Save Excel File", size_hint=(1, 0.16), background_color=(0.3, 0.7, 0.3, 1), font_size='16sp', bold=True
)
self.save_btn.bind(on_press=self.save_file)
center_box.add_widget(self.save_btn)

main_layout.add_widget(center_box)

# --- BOTTOM PANEL: MODE SELECTION ---
mode_row = BoxLayout(orientation="horizontal", size_hint=(1, 0.08), spacing=10)
mode_row.add_widget(Label(text="[b]Mode :[/b]", markup=True, size_hint=(0.3, 1), font_size='16sp'))

self.mode_spinner = Spinner(
text='Production', values=('Production', 'Dispatch'), size_hint=(0.7, 1), background_color=(0.7, 0.4, 0.2, 1), font_size='16sp'
)
mode_row.add_widget(self.mode_spinner)
main_layout.add_widget(mode_row)

self.message = Label(text="Status : Ready for Paste.", size_hint=(1, 0.07), color=(1, 1, 0, 1), font_size='14sp')
main_layout.add_widget(self.message)

Clock.schedule_once(self.lock_focus_on_input, 0.6)

return main_layout

def load_dispatch_locations_from_storage(self):
path_base = self.user_data_dir
self.locations_file_path = os.path.join(path_base, "dispatch_locations_data.json")
try:
if os.path.exists(self.locations_file_path):
with open(self.locations_file_path, 'r') as f: self.locations_list = json.load(f)
else: self.locations_list = ["Rolling Mill-1", "Billet Yard B", "Cooling Bed Yard"]
except Exception: self.locations_list = ["Rolling Mill-1", "Billet Yard B", "Cooling Bed Yard"]
if self.dispatch_location not in self.locations_list: self.locations_list.append(self.dispatch_location)

def update_storage_path(self):
if platform == 'android':
from android.storage import primary_external_storage_path
self.default_folder = os.path.join(primary_external_storage_path(), self.selected_location)
else:
self.default_folder = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop", self.selected_location)
if not os.path.exists(self.default_folder): os.makedirs(self.default_folder)

def lock_focus_on_input(self, dt):
if not self.is_popup_open:
if not self.barcode.focus: self.barcode.focus = True
Clock.schedule_once(self.lock_focus_on_input, 1.5)

def on_text_change(self, instance, value):
raw_data = value.strip()
current_time = time.time()
time_gap = current_time - self.last_input_time
self.last_input_time = current_time

if len(raw_data) > 1 and time_gap > 0.08: self.is_manual_entry = True
else: self.is_manual_entry = False

if raw_data and raw_data.count('/') >= 4:
self.barcode.unbind(text=self.on_text_change)
Clock.schedule_once(lambda dt: self.triple_check_and_process(self.barcode.text.strip()), 0.3)

def triple_check_and_process(self, final_data):
self.process_scan_data(final_data)
self.barcode.text = ""
self.barcode.bind(text=self.on_text_change)
Clock.schedule_once(lambda dt: setattr(self.barcode, 'focus', True), 0.05)

def process_scan_data(self, raw_text):
parts = raw_text.split("/")
current_mode = self.mode_spinner.text
entry_type_tag = "Manual" if self.is_manual_entry else "Scanned"

if len(parts) >= 5:
scanned_heat_raw = parts[0].strip().upper()
grade = parts[1].strip()
section = parts[2].strip()
length = parts[3].strip()
bloom = parts[4].strip()
 else: return

# 🛠️ SMART STORAGE: Live scan par heat ID me 'A' prefix auto-verify karega
if not scanned_heat_raw.startswith("A") and scanned_heat_raw.isdigit():
scanned_heat = "A" + scanned_heat_raw
else:
scanned_heat = scanned_heat_raw

if self.filename.text.strip() == "":
self.filename.text = scanned_heat

if scanned_heat.lower() != self.filename.text.strip().lower(): return

if current_mode == "Production":
duplicate_key = f"{scanned_heat}_{bloom}"
if duplicate_key in self.duplicates: return
self.duplicates.add(duplicate_key)
self.scan_data.append([scanned_heat, grade, section, length, bloom, "Produced", "CCM Cooling Bed", entry_type_tag, time.strftime("%Y-%m-%d %H:%M:%S")])
self.trigger_network_push(scanned_heat, bloom, "Produced", "CCM Cooling Bed")
else:
found = False
for row in self.scan_data:
# Excel loaded ya scanned values safety clean comparison check
row_heat_clean = str(row[0]).strip().upper()
if not row_heat_clean.startswith("A") and row_heat_clean.isdigit():
row_heat_clean = "A" + row_heat_clean

if row_heat_clean.lower() == scanned_heat.lower() and str(row[4]).strip().lower() == bloom.lower():
row[5] = "Dispatched"; row[6] = self.dispatch_location; row[7] = entry_type_tag; found = True
self.trigger_network_push(scanned_heat, bloom, "Dispatched", self.dispatch_location); break
if not found: return

self.count_label.text = f"No. of Count: {len(self.scan_data)}"
self.last_label.text = f"Last: {scanned_heat} / {bloom}"
self.refresh_history_ui()

def refresh_history_ui(self):
last_10_rows = self.scan_data[-10:]
for idx in range(10):
if idx < len(last_10_rows): self.history_labels[idx].text = f"{last_10_rows[idx][4]} -> {last_10_rows[idx][6]} [{last_10_rows[idx][7]}]"
else: self.history_labels[idx].text = ""

def clear_active_heat_data(self, instance):
self.scan_data.clear(); self.duplicates.clear(); self.filename.text = ""; self.barcode.text = ""
self.count_label.text = "No. of Count: 0"; self.last_label.text = "Last: None"; self.refresh_history_ui()
self.message.text = "Status: Fields cleared for next heat."

def undo_last_scan_row(self, instance):
if not self.scan_data: return
removed_item = self.scan_data.pop()
if f"{removed_item[0]}{removed_item[4]}" in self.duplicates: self.duplicates.remove(f"{removed_item[0]}{removed_item[4]}")
self.count_label.text = f"No. of Count: {len(self.scan_data)}"
self.last_label.text = f"Last Active: {self.scan_data[-1][0]}" if self.scan_data else "Last: None"
self.refresh_history_ui()

def open_production_report_popup(self, instance):
self.is_popup_open = True

form_layout = BoxLayout(orientation="vertical", padding=12, spacing=10)
grid_inputs = GridLayout(cols=2, padding=5, spacing=10, size_hint_y=0.8)

grid_inputs.add_widget(Label(text="📍 Report Source:", font_size='14sp', bold=True))
source_spinner = Spinner(text=self.rep_source_mode, values=("1. Use Current Live Scan", "2. Extract From Saved Folder File"), font_size='13sp')
grid_inputs.add_widget(source_spinner)

grid_inputs.add_widget(Label(text="🔑 Heat Number:", font_size='14sp'))
heat_in = TextInput(text=self.filename.text if self.filename.text else "", multiline=False, font_size='14sp')
grid_inputs.add_widget(heat_in)

grid_inputs.add_widget(Label(text="🏢 Customer Name:", font_size='14sp'))
customer_in = TextInput(text=self.rep_cust_name, hint_text="e.g. Shivom / G.N.A", multiline=False, font_size='14sp')
grid_inputs.add_widget(customer_in)

grid_inputs.add_widget(Label(text="✂️ Front End Discard (mm):", font_size='14sp'))
front_discard_in = TextInput(text=self.rep_front_discard, multiline=False, font_size='14sp')
grid_inputs.add_widget(front_discard_in)

grid_inputs.add_widget(Label(text="✂️ End Cutting Discard (mm):", font_size='14sp'))
end_discard_in = TextInput(text=self.rep_end_discard, multiline=False, font_size='14sp')
grid_inputs.add_widget(end_discard_in)

grid_inputs.add_widget(Label(text="❄️ Cooling System Type:", font_size='14sp'))
cooling_in = TextInput(text=self.rep_cooling_type, hint_text="e.g. Fast Fan Cooling", multiline=False, font_size='14sp')
grid_inputs.add_widget(cooling_in)

form_layout.add_widget(grid_inputs)

btn_container = BoxLayout(orientation="horizontal", spacing=10, size_hint_y=0.2)
generate_btn = Button(text="📲 Generate & Send WhatsApp", background_color=(0.1, 0.6, 0.2, 1), bold=True)
cancel_btn = Button(text="Cancel", background_color=(0.7, 0.2, 0.2, 1))
btn_container.add_widget(generate_btn); btn_container.add_widget(cancel_btn)
form_layout.add_widget(btn_container)

config_popup = Popup(title="📊 Billet Production Slip Configurator", content=form_layout, size_hint=(0.95, 0.9), auto_dismiss=False)

def process_and_blast_whatsapp(btn_instance):
self.rep_cust_name = customer_in.text.strip()
self.rep_front_discard = front_discard_in.text.strip()
self.rep_end_discard = end_discard_in.text.strip()
self.rep_cooling_type = cooling_in.text.strip()
self.rep_source_mode = source_spinner.text

# 🛠️ SMART MATCH INTEGRATION FIXED: Auto-capital sanitation matching logic
input_heat = heat_in.text.strip().upper()
if input_heat == "": return

if not input_heat.startswith("A") and input_heat.isdigit():
target_heat = "A" + input_heat
else:
target_heat = input_heat

data_source_list = []
if "Live Scan" in self.rep_source_mode:
if not self.scan_data: return
data_source_list = self.scan_data
else:
full_path = os.path.join(self.default_folder, target_heat + self.selected_file_type)
if not os.path.exists(full_path): return
try:
wb = load_workbook(full_path); ws = wb.active; is_header = True
for row in ws.iter_rows(values_only=True):
if not row or row[0] is None: continue
if is_header and str(row[0]).strip().lower() == "heat no": is_header = False; continue
data_source_list.append(list(row))
except Exception: return

if not data_source_list: return

try:
s1_regular_lengths = []
s2_regular_lengths = []
s1_al_billets = []
s2_bl_billets = []
all_lengths = []
grade = "20MnCr5"
display_section = "160X160"
total_raw_weight_kg = 0.0

weight_per_mtr_map = {"130x130": 121.0, "160x160": 200.0, "200x200": 314.0, "250x250": 498.0, "320x250": 628.0}

for r_idx, row in enumerate(data_source_list):
grade = str(row[1])
excel_section = str(row[2]).strip().lower().replace(" ", "")
display_section = str(row[2]).strip()

raw_len_str = str(row[3]).strip().upper().replace("M", "").replace("R", "").strip()
row_length = float(raw_len_str)
bloom_id = str(row[4]).strip().upper()

all_lengths.append(row_length)

if "AL" in bloom_id or bloom_id == "AL":
s1_al_billets.append(row_length)
elif "BL" in bloom_id or bloom_id == "BL":
s2_bl_billets.append(row_length)
else:
if r_idx % 2 == 0: s1_regular_lengths.append(row_length)
else: s2_regular_lengths.append(row_length)

wt_per_mtr = weight_per_mtr_map.get(excel_section, 200.0)
total_raw_weight_kg += wt_per_mtr * row_length

length_counts = Counter(all_lengths)
standard_length = length_counts.most_common(1)[0][0]

s1_counts_map = Counter(s1_regular_lengths)
s2_counts_map = Counter(s2_regular_lengths)

s1_lines = []
for len_val, count in s1_counts_map.items():
s1_lines.append(f"👉 {len_val} Mtr — {count} Pcs")
for len_val in s1_al_billets:
s1_lines.append(f"👉 {len_val} Mtr — AL")
s1_display_block = "\n".join(s1_lines) if s1_lines else "👉 None"

s2_lines = []
for len_val, count in s2_counts_map.items():
s2_lines.append(f"👉 {len_val} Mtr — {count} Pcs")
for len_val in s2_bl_billets:
s2_lines.append(f"👉 {len_val} Mtr — BL")
s2_display_block = "\n".join(s2_lines) if s2_lines else "👉 None"

remainder_kg = total_raw_weight_kg % 100.0
if remainder_kg >= 50.0:
rounded_weight_mt = (int(total_raw_weight_kg / 100.0) * 100.0 + 100.0) / 1000.0
else:
rounded_weight_mt = (int(total_raw_weight_kg / 100.0) * 100.0) / 1000.0

final_total_wt = round(rounded_weight_mt, 2)
total_billets = len(data_source_list)

final_slip_report = (
f"------------- ARORA IRON & STEEL R/M ------------- \n"
f" Billet Production Slip \n"
f"---------------------------------------------\n"
f"🔷 Heat No: {target_heat}\n"
f"🔷 Customer Name: {self.rep_cust_name}\n"
f"🔷 Grade: {grade}\n"
f"🔷 Section: {display_section}\n"
f"🔷 Total WT: {final_total_wt:0.2f} MT\n"
f"🔷 Length: {standard_length} Mtr\n"
f"---------------------------------------------\n"
f"🔷 No Of Billet S1:\n{s1_display_block}\n"
f"---------------------------------------------\n"
f"🔷 No Of Billet S2:\n{s2_display_block}\n"
f"---------------------------------------------\n"
f"🔷 Front End Discard: {self.rep_front_discard} MM\n"
f"🔷 End Cutting Discard: {self.rep_end_discard} MM\n"
f"🔷 Total Billet: {total_billets}\n"
f"🔷 Cooling Type: {self.rep_cooling_type}\n"
f"---------------------------------------------\n"
f"Generated on: {time.strftime('%Y-%m-%d %H:%M:%S')}\n"
)

try:
target_filename = f"Slip_{target_heat}.txt"
txt_full_path = os.path.join(self.default_folder, target_filename)
with open(txt_full_path, "w", encoding="utf-8") as text_file:
text_file.write(final_slip_report)
except Exception: pass

encoded_msg = urllib.parse.quote(final_slip_report)
webbrowser.open(f"https://wa.me/?text={encoded_msg}")

self.is_popup_open = False
config_popup.dismiss()

except Exception:
self.is_popup_open = False
config_popup.dismiss()

def close_form_and_release_focus(btn_ins):
self.is_popup_open = False
config_popup.dismiss()

generate_btn.bind(on_release=process_and_blast_whatsapp)
cancel_btn.bind(on_release=close_form_and_release_focus)
config_popup.open()

def trigger_network_push(self, heat, bloom, status, location):
if not self.enable_ip_push or not requests: return
def push_thread():
url = f"http://{self.selected_ip}:{self.selected_port}/api/scan"
try: requests.post(url, json={"heat": heat, "bloom": bloom, "status": status, "location": location}, timeout=2)
except Exception: pass
threading.Thread(target=push_thread, daemon=True).start()

def save_file(self, instance):
filename = self.filename.text.strip()
if filename == "" or not self.scan_data: return
full_path = os.path.join(self.default_folder, filename + self.selected_file_type)
try:
if os.path.exists(full_path):
wb = load_workbook(full_path); ws = wb.active; ws.protection.sheet = False; ws.delete_rows(1, ws.max_row+1)
else: wb = Workbook(); ws = wb.active
ws.append(["Heat No", "Grade", "Section", "Length", "Billet ID / Bloom", "Status", "Current Location", "Entry Type Tag", "Timestamp"])
for row in self.scan_data: ws.append(row)

ws.protection.sheet = True
ws.protection.password = "PPC_CCM_Secure_2026"
ws.protection.enable()

wb.save(full_path)
self.message.text = f"Status : Excel Saved & Locked! -> {filename}"
except Exception as e: self.message.text = f"Status : Save Error! {str(e)}"

def open_existing_file(self, instance):
filename_raw = self.filename.text.strip().upper()
if filename_raw == "": return

if not filename_raw.startswith("A") and filename_raw.isdigit():
filename = "A" + filename_raw
else:
filename = filename_raw

full_path = os.path.join(self.default_folder, filename + self.selected_file_type)
if not os.path.exists(full_path): return
try:
wb = load_workbook(full_path); ws = wb.active; self.scan_data.clear(); self.duplicates.clear()
is_header = True
for row in ws.iter_rows(values_only=True):
if not row or row[0] is None: continue
if is_header and str(row[0]).strip().lower() == "heat no": is_header = False; continue
row_list = list(row)
while len(row_list) < 9: row_list.append("")
self.scan_data.append(row_list)
self.duplicates.add(f"{row_list[0]}_{row_list[4]}")
if self.scan_data:
self.filename.text = str(self.scan_data[-1][0])
self.last_label.text = f"Loaded: {self.scan_data[-1][0]}/{self.scan_data[-1][4]}"
self.count_label.text = f"No. of Count: {len(self.scan_data)}"; self.refresh_history_ui()
except Exception as e: self.message.text = f"Status : Error opening file: {str(e)}"

def open_settings_popup(self, instance):
self.is_popup_open = True
content = GridLayout(cols=2, padding=10, spacing=10)

content.add_widget(Label(text="Excel Save Folder:"))
folder_in = TextInput(text=self.selected_location, multiline=False)
content.add_widget(folder_in)

content.add_widget(Label(text="Dispatch Target Yard:", font_size='14sp', bold=True))
self.location_spinner = Spinner(text=self.dispatch_location, values=tuple(self.locations_list), font_size='14sp')
content.add_widget(self.location_spinner)

content.add_widget(Label(text="Add New Yard Location:", font_size='14sp'))
new_loc_input = TextInput(hint_text="Type new yard name...", multiline=False, font_size='14sp')
content.add_widget(new_loc_input)

content.add_widget(Label(text="Central PC IP:"))
ip_in = TextInput(text=self.selected_ip, multiline=False)
content.add_widget(ip_in)

content.add_widget(Label(text="Server Port:"))
port_in = TextInput(text=self.selected_port, multiline=False)
content.add_widget(port_in)

btn_row = BoxLayout(orientation="horizontal", spacing=10, size_hint_y=None, height=40)
save_btn = Button(text="Save Settings", background_color=(0.2, 0.7, 0.2, 1), bold=True)
cancel_btn = Button(text="Cancel", background_color=(0.7, 0.2, 0.2, 1))
btn_row.add_widget(save_btn); btn_row.add_widget(cancel_btn)

popup = Popup(title="⚙️ Settings Center", content=BoxLayout(orientation="vertical", spacing=10), size_hint=(0.9, 0.8), auto_dismiss=False)
popup.content.add_widget(content)
popup.content.add_widget(btn_row)

def save_and_close(btn_instance):
self.selected_location = folder_in.text.strip()

typed_new_location = new_loc_input.text.strip()
if typed_new_location != "":
if typed_new_location not in self.locations_list:
self.locations_list.append(typed_new_location)
try:
with open(self.locations_file_path, 'w') as f: json.dump(self.locations_list, f)
except Exception: pass
self.dispatch_location = typed_new_location
else:
self.dispatch_location = self.location_spinner.text

self.selected_ip = ip_in.text.strip()
self.selected_port = port_in.text.strip()
self.update_storage_path()

self.is_popup_open = False
popup.dismiss()
self.message.text = f"Status : Config synced! Active Target: '{self.dispatch_location}'"

def cancel_settings(btn_instance):
self.is_popup_open = False
popup.dismiss()

save_btn.bind(on_release=save_and_close)
cancel_btn.bind(on_release=cancel_settings)
popup.open()

if name == "main":
ScannerApp().run()